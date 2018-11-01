import openpyxl
from openpyxl import Workbook
filepath = "/home/viki/Python/Scripts/trial.xlsx"
wb = openpyxl.Workbook()

#wb_obj = openpyxl.load_workbook(filepath)
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb.active

# Cell objects also have row, column,  
# and coordinate attributes that provide 
# location information for the cell. 

# Note: The first row or  
# column integer is 1, not 0. 

# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
#sheet['A1'] = 1
sheet_obj.cell(row=2, column=2).value = 4
#sheet.cell(row=2, column=2).value = 2


wb.save(filepath)
